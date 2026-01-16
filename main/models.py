import csv
import hashlib
import uuid

import openpyxl
from django.contrib.auth.models import AbstractUser
from django.core.validators import FileExtensionValidator
from django.db import models
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.utils.translation import gettext_lazy as _

from .custom_functions import autotask
from .custom_validators import validate_emails_in_file


@autotask
def send_password(email_to, username, password):
    pass
    # subject = 'EasyTP by KuberLeads'

    # message = render_to_string("main/send_password.html", {'username': username, 'password': password})
    # send_mail(subject=subject, message=message, from_email=settings.EMAIL_HOST_USER, recipient_list=[email_to],
    #           fail_silently=False)


class AccessGroup(models.Model):
    FULL = "Full Access Group"
    CP1 = "Cycle Préparatoire 1"
    CP2 = "Cycle Préparatoire 2"
    CS1 = "Second Cycle 1"
    CS2 = "Second Cycle 2"
    GUEST = "Guest"

    GROUPS = [
        (FULL, "Full Access Group"),
        (CP1, "Cycle Préparatoire 1"),
        (CP2, "Cycle Préparatoire 2"),
        (CS1, "Second Cycle 1"),
        (CS2, "Second Cycle 2"),
        (GUEST, "Guest"),
    ]

    name = models.CharField(max_length=25, default=CP1, unique=True, blank=False)

    def __str__(self):
        return f"{self.name}"

    def has_access_to(self):
        if self.name == self.FULL:
            return "All Apps"
        return ", ".join([app.name for app in self.apps.all()])

    def get_apps(self):
        return [app.name for app in self.apps.all()]


class App(models.Model):
    name = models.CharField(max_length=50, blank=False, unique=True)
    group = models.ManyToManyField(AccessGroup, related_name="apps", blank=True)
    image = models.CharField(max_length=100)

    def __str__(self):
        return f"{self.name}"

    def groups(self):
        return ", ".join([g.name for g in self.group.all()])


class DefaultUser(AbstractUser):
    email = models.EmailField(
        _("email address"),
        max_length=50,
        blank=False,
        unique=True,
        help_text=_("Enter the email of user"),
        #   validators=[EsiEmailValidator(allowlist=['esi.dz'],
        #                           message='Enter a valid "@esi.dz" email address.')],
    )

    GUEST = "G"
    TEACHER = "T"
    STUDENT = "S"
    ADMIN = "A"
    ROLES = [
        (TEACHER, "Teacher"),
        (STUDENT, "Student"),
        (ADMIN, "Staff"),
        (GUEST, "Guest"),
    ]

    role = models.CharField(max_length=1, choices=ROLES, default=STUDENT, blank=False)

    group = models.ForeignKey(
        AccessGroup, on_delete=models.SET_DEFAULT, default=None, null=True, related_name="students"
    )

    upload_limit = models.IntegerField(default=50, help_text=_("Upload limit in MB for the user"))

    size_uploaded = models.IntegerField(
        default=0, help_text=_("Total size uploaded by the user in MB")
    )

    def is_guest(self):
        return self.role == self.GUEST

    def save(self, *args, **kwargs):
        # self.username = self.email.split('@')[0].replace('.', '-').replace('_', '-').replace('+', '-')

        if self.role == self.ADMIN or self.is_superuser:
            self.is_staff = True

            if self.is_superuser:
                self.role = self.ADMIN

            # add superusers,staff to FULL ACCESS GROUP

            self.group = AccessGroup.objects.get_or_create(name=AccessGroup.FULL)[0]

        super().save(*args, **kwargs)

    def apps_available(self):
        if not self.group:
            return "Not in a group yet"
        return self.group.has_access_to()


class UsersFromCSV(models.Model):
    file = models.FileField(
        default="", validators=[FileExtensionValidator(["csv", "xlsx"]), validate_emails_in_file]
    )

    role = models.CharField(
        max_length=1, choices=DefaultUser.ROLES, default=DefaultUser.STUDENT, blank=False
    )
    group = models.ForeignKey(AccessGroup, on_delete=models.SET_DEFAULT, default=None)

    def save_user(self, email, last_name, first_name):
        if email:
            user_exist = DefaultUser.objects.filter(email=email)
            if user_exist:
                try:
                    user_exist.update(
                        email=email,
                        role=self.role,
                        group=self.group,
                        last_name=last_name,
                        first_name=first_name,
                    )
                except Exception as e:
                    print("user ", email, " not updated")
                    print(e)
            else:
                try:
                    username = email.split("@")[0]
                    password = uuid.uuid4().hex[:8]

                    DefaultUser.objects.create_user(
                        email=email,
                        password=password,
                        role=self.role,
                        group=self.group,
                        username=email.split("@")[0],
                        last_name=last_name,
                        first_name=first_name,
                    )
                    try:
                        send_password(email, username, password)
                    except Exception as e:
                        print("failed to send email", e)

                except Exception as e:
                    print("user ", email, "not created")
                    print(e)

    def save(self, force_insert=False, force_update=False, using=None, update_fields=None):
        if str(self.file).endswith(".csv"):
            data = csv.reader(self.file)
            next(data)  # Skip header row

            for row in data:
                self.save_user(email=row[0], last_name=row[1], first_name=row[2])

        else:
            sheet = openpyxl.load_workbook(self.file)
            sheet = sheet.active

            for row in range(2, sheet.max_row + 1):
                self.save_user(
                    email=sheet.cell(row=row, column=1).value,
                    last_name=sheet.cell(row=row, column=2).value,
                    first_name=sheet.cell(row=row, column=3).value,
                )

    def __str__(self):
        return self.role + "s"


class Pod(models.Model):
    pod_user = models.ForeignKey(
        DefaultUser, on_delete=models.CASCADE, default=None, related_name="pod_user"
    )
    app_name = models.CharField(max_length=200, default=None, blank=False, null=True)
    pod_name = models.CharField(max_length=200, default=None, blank=False, null=True)
    # pod_port = models.CharField(max_length=200, default=None, blank=False, null=True)
    pod_vnc_user = models.CharField(max_length=200, default=None, blank=False, null=True)
    pod_vnc_password = models.CharField(max_length=200, default=None, blank=False, null=True)

    date_created = models.DateTimeField(auto_now_add=True, blank=True)
    date_modified = models.DateTimeField(auto_now=True, blank=True)
    pod_namespace = models.CharField(max_length=200, default=None, blank=False, null=True)

    is_deployed = models.BooleanField(default=False, help_text=_("Is the pod deployed?"))

    def __str__(self):
        return f"{self.pod_user.username}:{self.pod_name}:{self.app_name}"


@receiver(post_save, sender=DefaultUser)
def generate_pods(sender, instance, created, **kwargs):
    if created:
        # Check if user has a group assigned before trying to access apps
        if not instance.group:
            # Log this for debugging - user was created without a group
            import logging

            logger = logging.getLogger(__name__)
            logger.warning(
                f"User {instance.username} created without group assignment. No pods will be generated."
            )
            return

        for app in instance.group.apps.all():
            app_name = app.name
            pod = Pod(
                pod_user=instance,
                app_name=app_name,
                pod_name=hashlib.md5(
                    f"{app_name}:{instance.username}:{instance.id}".encode("utf-8")
                ).hexdigest(),
                pod_vnc_user=uuid.uuid4().hex[:6],
                pod_vnc_password=uuid.uuid4().hex,
                pod_namespace="apps",
            )
            pod.save()


class Instances(models.Model):
    pod = models.OneToOneField(Pod, on_delete=models.CASCADE, default=None, related_name="instance")
    instance_name = models.CharField(max_length=200, default=None, blank=False, null=True)
    date_created = models.DateTimeField(auto_now_add=True, blank=True)
    date_modified = models.DateTimeField(auto_now=True, blank=True)
    novnc_url = models.URLField(blank=True, null=True)

    def __str__(self):
        return f"{self.pod}:{self.instance_name}"


class UserActivity(models.Model):
    """Model to track user activities"""

    # Activity types
    LOGIN = "login"
    LOGOUT = "logout"
    POD_START = "pod_start"
    POD_STOP = "pod_stop"
    FILE_UPLOAD = "file_upload"
    FILE_DELETE = "file_delete"
    FILE_DOWNLOAD = "file_download"
    ACCOUNT_CREATED = "account_created"
    PAGE_VIEW = "page_view"

    ACTIVITY_CHOICES = [
        (LOGIN, "Login"),
        (LOGOUT, "Logout"),
        (POD_START, "Pod Started"),
        (POD_STOP, "Pod Stopped"),
        (FILE_UPLOAD, "File Uploaded"),
        (FILE_DELETE, "File Deleted"),
        (FILE_DOWNLOAD, "File Downloaded"),
        (ACCOUNT_CREATED, "Account Created"),
        (PAGE_VIEW, "Page View"),
    ]

    user = models.ForeignKey(
        "DefaultUser", on_delete=models.SET_NULL, related_name="activities", null=True, blank=True
    )
    username = models.CharField(max_length=150, blank=True)  # Store username separately
    activity_type = models.CharField(max_length=20, choices=ACTIVITY_CHOICES)
    timestamp = models.DateTimeField(auto_now_add=True)
    ip_address = models.GenericIPAddressField(null=True, blank=True)
    user_agent = models.TextField(null=True, blank=True)
    details = models.JSONField(default=dict, blank=True)  # Store additional context

    class Meta:
        ordering = ["-timestamp"]
        verbose_name = "User Activity"
        verbose_name_plural = "User Activities"
        indexes = [
            models.Index(fields=["-timestamp"]),
            models.Index(fields=["user", "-timestamp"]),
            models.Index(fields=["username", "-timestamp"]),  # Add index for username
            models.Index(fields=["activity_type", "-timestamp"]),
        ]

    def __str__(self):
        username = self.username or (self.user.username if self.user else "Unknown")
        return f"{username} - {self.get_activity_type_display()} - {self.timestamp}"
